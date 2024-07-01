import os
from openpyxl import load_workbook
import re
from tqdm import tqdm


def process_excel_file(file_path, replace_values):
    try:
        workbook = load_workbook(file_path)
        replacements = 0

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for pattern, replacement in replace_values:
                            replaced_text, count = re.subn(
                                pattern, replacement, cell.value)
                            if count > 0:
                                cell.value = replaced_text
                                replacements += count
        workbook.save(file_path)

        if replacements > 0:
            return f"Fayl muvaffaqiyatli o'zgartirildi 🟩: {file_path}\nFayldagi almashtirishlar soni: {replacements} ta"
        else:
            return f"Fayl o'zgartirilmadi 🟨: {file_path}\nFayldagi almashtirishlar yo'q ❌"
    except Exception as e:
        return f"Faylni qayta ishlashda xatolik 🟥: {file_path}. Sabab => : {str(e)}"


def main():
    log_file = open("log.txt", "w", encoding="utf-8")
    folder_path = input(
        "Iltimos, .xlsx fayllari joylashgan papka manzilini kiriting 📂: ")

    # Заданный словарь замен
    replace_values = [
        ("о", "o"), ("О", "O"),
        ("А", "A"), ("а", "a"),
        ("е", "e"), ("E", "E"),
        ("М", "M"), ("м", "m"),
        ("С", "C"), ("с", "c"),
        ("Х", "X"), ("х", "x"),
        ("К", "K"), ("к", "k"),
        ("Т", "T"), ("т", "t"),
        ("ы", "i"),
        ("`", "'"), #("ʼ", "'"),
        ("ʻ", "'"), ("ʼ", "'"),
        ("‘", "'"), ("’", "'")
        # ("aim", "ayim"), ("aid", "ayid"),
        # ("oev", "oyev"), ("aev", "ayev"),
        # ("iev", "iyev"), ("bdi", "bdu"),
        # ("oi", "oyi"),
        # ("itdi", "iddi"), ("bduy", "bdiy"),
        # ("ulloh", "ullo"), ("illo", "ullo"),
        # ("ullo", "ulla"), ("atulla", "atilla"),
        # ("allah", "ulloh"), ("adulla", "odulla"),
        # ("ulloh", "ullo"), ("Sodulla", "Sadulla"),
        # ("xuja", "xo'ja"), ("hoja", "xo'ja"),
        # ("ho'ja", "xo'ja"), ("kizi", "qizi"),
        # ("ugli", "o'g'li"), ("ug'li", "o'g'li"),
        # ("o'gli", "o'g'li"), (" uli", "o'g'li"),
        # ("og'li", "o'g'li"), ("ogli", "o'g'li"),
        # (" uli", " o'g'li"), (" o'li", "o'g'li"),
        # (" ogli", "o'g'li"),
        

        # #step2
        # ("ovna", " qizi"), ("ivch", " vich"),
        # ("yevna", " qizi"), ("evna", " qizi"),
        # ("evich", " o'g'li"), ("yevich", " o'g'li"),
        # ("ovich", " o'g'li"), ("-", " "),
        
        # #step3
        # ("boy", ""), ("jon", ""),
        # ("gir", ""), ("bek", ""),
        # ("xo'ja", ""), ("xon", ""),
        # ("begim", ""), ("iddin", ""),
        # ("itdin", ""), ("utdin", ""),
        # ("uddin", ""), ("nazar", ""),
        # ("momo", "mono"), ("sroch", "siroj"),
        # ("sroj", "siroj"), ("ukr", "ukur"),
        # ("ehzod", "egzod"), ("ekzod", "egzod"),
        # ("ehzod", "egzod"), ("Jaxon", "Jahon"),
        # ("nazar", ""),

        # #step4
        # ("O'", "U"), ("o'", "u"),
        # ("U", "I"), ("u", "i"),
        # ("G'", "G"), ("g'", "g"),
        # ("H", "X"), ("h", "x"),
        # ("Q", "K"), ("q", "k"),
        # ("st", "s"), ("ts", "s"),

        # #step6
        # ("O", "A"), ("o", "a"),
        # ("Dj", "J"), ("dj", "j"),
        # ("Cx", "Ch"), ("cx", "ch"),
        # ("Sx", "Sh"), ("sx", "sh"),
        
        # #step5
        # ("bay", ""), ("jan", ""),
        # ("xan", ""), ("axxa", "axa"), # !!!        

        # # warn step:
        # ("m", "n"), ("zz", "z"),
        # ("d", "t"), ("xx", "x"),
        # ("Ye", "E"), # ("ye", "e"), 
        # ("ll", "l"), ("tt", "t"),
        # ("aa", "a"), ("vv", "v"),
        # ("ff", "f"), ("bb", "b"),
        # ("mm", "m"),

        # # Critical:
        # ("E", "I"), ("e", "i"),

        # #step7
        # ("'", "")
    ]

    files = [os.path.join(root, file_name) for root, _, file_names in os.walk(
        folder_path) for file_name in file_names if file_name.endswith('.xlsx')]
    success_count = 0
    fail_count = 0
    log = []

    for file in tqdm(files, desc="FAYLLAR QAYTA ISHLASH JARAYONIDA 🔁:"):
        log_entry = process_excel_file(file, replace_values)
        log.append(log_entry)
        if "muvaffaqiyatli" in log_entry.lower():
            success_count += 1
        else:
            fail_count += 1

    for entry in log:
        log_file.write(entry + "\n")
        print(entry)

    log_file.close()
    print("✅✅✅ OPERATSIYA MUVAFAQQIYATLI YAKUNLANDI ✅✅✅")
    print(f"❇️ Muvafaqqiyatli operatsiyalar soni: {success_count}")
    print(f"❌ Muvafaqqiyatsiz operatsiyalar soni: {fail_count}")


if __name__ == "__main__":
    main()