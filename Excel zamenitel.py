import os
from openpyxl import load_workbook
import re

# Papkaga yo'l:
folder_path = input("Iltimos, .xlsx fayllari joylashgan papka manzilini kiriting: ")

# Almashtirish:
replace_values = {}
while True:
    key = input("O'zgartirilishi kerak bo'lgan matnini kiriting (tugatish uchun bo'sh qoldiring) 🔍: ")
    if not key:
        break
    value = input("O'zgartirish uchun yangi qiymatni kiriting 📝: ")
    replace_values[key] = value

if not replace_values:
    print("O'zgartirish qiymatlari kiritilmadi. Skript yakunlandi 🥴")
else:
    confirmation = input("O'zgartirishni boshlash uchun [HA] ni, bekor qilish uchun [YO'Q] ni kiriting")
    if confirmation.lower() == 'HA':

        # O'zgartirishlarni kuzatish uchun log
        log = []

        # Excel fayl obrabotka funksiyasi
        def process_excel_file(file_path):
            try:
                workbook = load_workbook(file_path)

                replacements = 0  # Fayldagi almashtinishlar soni

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value is not None and isinstance(cell.value, str):
                                for pattern, replacement in replace_values.items():
                                    replaced_text, count = re.subn(pattern, replacement, cell.value)
                                    if count > 0:
                                        cell.value = replaced_text
                                        replacements += count

                workbook.save(file_path)
                if replacements > 0:
                    log_entry = f"Файл успешно изменен 🟩: {file_path}\nКоличество замен в файле: {replacements}"
                    log.append(log_entry)
                else:
                    log_entry = f"Файл не изменен 🟨: {file_path}\nКоличество замен в файле: 0 ❌"
                    log.append(log_entry)
            except Exception as e:
                log_entry = f"Ошибка обработки файла 🟥: {file_path}. Причина => : {str(e)}"
                log.append(log_entry)

        # Fayl va papkalarni qayta ishlash uchun rekursiv funksiya
        def process_folder(folder_path):
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.endswith('.xlsx'):
                        file_path = os.path.join(root, file)
                        process_excel_file(file_path)

        # Papkani qayta obrabotka funksiyasini chaqirish
        process_folder(folder_path)

        # Logni chaqirish
        for entry in log:
            print(entry)

        print(" _______           _ _                                                       ")
        print("(_______)         | | |                                                      ")
        print(" _____ _____ _   _| | | _____  ____                                          ")
        print("|  ___|____ | | | | | |(____ |/ ___)                                         ")
        print("| |   / ___ | |_| | | |/ ___ | |                                             ")
        print("|_|   \_____|\__  |\_)_)_____|_|                                             ")
        print("            (____/                                                           ")
        print("                            ___                   _                   _  _   ")
        print("                           / __)                 (_)              _  | |(_)  ")
        print("____   _   _ _   _ _____ _| |__ _____  ____  ____ _ _   _ _____ _| |_| | _   ")
        print("|    \| | | | | | (____ (_   __|____ |/ _  |/ _  | | | | (____ (_   _) || |  ")
        print("| | | | |_| |\ V // ___ | | |  / ___ | |_| | |_| | | |_| / ___ | | |_| || |  ")
        print("|_|_|_|____/  \_/ \_____| |_|  \_____|\__  |\__  |_|\__  \_____|  \__)\_)_|  ")
        print("                                         |_|   |_| (____/                    ")
        print("     _                             _       _ _     _ _     🟩🟩🟩🟩🟩🟩🟩 ")
        print("    ( )                        _  (_)     (_) |   | (_)    🟩🟩🟩🟩🟩◽️🟩 ")
        print("  __|/_____ ____ _____  ____ _| |_ _  ____ _| | __| |_     🟩🟩🟩🟩◽️◽️🟩 ")
        print(" / _ (___  ) _  (____ |/ ___|_   _) |/ ___) | |/ _  | |    🟩◽️🟩◽️◽️🟩🟩 ")
        print("| |_| / __( (_| / ___ | |     | |_| | |   | | ( (_| | |    🟩◽️◽️◽️🟩🟩🟩 ")
        print(" \___(_____)___ \_____|_|      \__)_|_|   |_|\_)____|_|    🟩🟩◽️🟩🟩🟩🟩 ")
        print("            __| |                                          🟩🟩🟩🟩🟩🟩🟩 ")
        print("           (___/                                                             ")
    else:
        print("Замена отменена :(")